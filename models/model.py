import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
import numpy as np
import os
import time

from collections import defaultdict
from openpyxl import Workbook
from torchvision import datasets, models

def get_model(model_dir, device, mtype='OVR'):
    # Create the loss function
    criterion = nn.CrossEntropyLoss()
    
    # Create the covid binary model
    covid_ft = models.resnet50(pretrained=True)
    covid_num_ftrs = covid_ft.fc.in_features
    covid_ft.fc = nn.Linear(covid_num_ftrs, 2)

    # 473
    covid_ft.load_state_dict(torch.load(f"{model_dir}/473/best_covid_binary.pt"))
    # # 517
    # covid_ft.load_state_dict(torch.load(f"{model_dir}/covid_binary2.pt"))
    # covid_ft.load_state_dict(torch.load(f"{model_dir}/517/final_covid_binary.pt"))
    # covid_ft.load_state_dict(torch.load(f"{model_dir}/basic_covid_binary.pt"))

    covid_ft = covid_ft.to(device)
    
    if mtype == 'OVR':
        # Create the normal binary model
        normal_ft = models.resnet50(pretrained=True)
        normal_num_ftrs = normal_ft.fc.in_features
        normal_ft.fc = nn.Linear(normal_num_ftrs, 2)

        # 473
        normal_ft.load_state_dict(torch.load(f"{model_dir}/473/best_normal_binary.pt"))
        # # 517
        # normal_ft.load_state_dict(torch.load(f"{model_dir}/normal_binary2.pt"))
        # normal_ft.load_state_dict(torch.load(f"{model_dir}/517/final_normal_binary.pt"))
        # normal_ft.load_state_dict(torch.load(f"{model_dir}/basic_normal_binary.pt"))
    
        normal_ft = normal_ft.to(device)

        # Create the pneumonia binary model
        pneumonia_ft = models.resnet50(pretrained=True)
        pneumonia_num_ftrs = pneumonia_ft.fc.in_features
        pneumonia_ft.fc = nn.Linear(pneumonia_num_ftrs, 2)

        # 473
        pneumonia_ft.load_state_dict(torch.load(f"{model_dir}/473/best_pneumonia_binary.pt"))
        # # 517
        # pneumonia_ft.load_state_dict(torch.load(f"{model_dir}/pneumonia_binary2.pt"))
        # pneumonia_ft.load_state_dict(torch.load(f"{model_dir}/517/final_pneumonia_binary.pt"))
        # pneumonia_ft.load_state_dict(torch.load(f"{model_dir}/basic_pneumonia_binary.pt"))

        pneumonia_ft = pneumonia_ft.to(device)
        
        cov_nor_ft = models.resnet50(pretrained=True)
        cov_nor_num_ftrs = cov_nor_ft.fc.in_features
        cov_nor_ft.fc = nn.Linear(cov_nor_num_ftrs, 2)

        # 473
        cov_nor_ft.load_state_dict(torch.load(f"{model_dir}/473/best_normal_pneumonia_binary.pt"))
        # # 517
        # cov_nor_ft.load_state_dict(torch.load(f"{model_dir}/covid_normal_binary.pt"))
        # cov_nor_ft.load_state_dict(torch.load(f"{model_dir}/517/final_covid_normal_binary.pt"))
        # cov_nor_ft.load_state_dict(torch.load(f"{model_dir}/basic_covid_normal_binary.pt"))
        cov_nor_ft = cov_nor_ft.to(device)
        
        cov_pneu_ft = models.resnet50(pretrained=True)
        cov_pneu_num_ftrs = cov_pneu_ft.fc.in_features
        cov_pneu_ft.fc = nn.Linear(cov_pneu_num_ftrs, 2)

        # 473
        cov_pneu_ft.load_state_dict(torch.load(f"{model_dir}/473/best_covid_pneumonia_binary.pt"))
        # # 517
        # cov_pneu_ft.load_state_dict(torch.load(f"{model_dir}/covid_pneumonia_binary.pt"))
        # cov_pneu_ft.load_state_dict(torch.load(f"{model_dir}/517/final_covid_pneumonia_binary.pt"))
        # cov_pneu_ft.load_state_dict(torch.load(f"{model_dir}/basic_covid_pneumonia_binary.pt"))
        cov_pneu_ft = cov_pneu_ft.to(device)
        
        nor_pneu_ft = models.resnet50(pretrained=True)
        nor_pneu_num_ftrs = nor_pneu_ft.fc.in_features
        nor_pneu_ft.fc = nn.Linear(nor_pneu_num_ftrs, 2)

        # 473
        nor_pneu_ft.load_state_dict(torch.load(f"{model_dir}/473/best_normal_pneumonia_binary.pt"))
        # # 517
        # nor_pneu_ft.load_state_dict(torch.load(f"{model_dir}/normal_pneumonia_binary.pt"))
        # nor_pneu_ft.load_state_dict(torch.load(f"{model_dir}/517/final_pneumonia_normal_binary.pt"))
        # nor_pneu_ft.load_state_dict(torch.load(f"{model_dir}/basic_pneumonia_normal_binary.pt"))
        nor_pneu_ft = nor_pneu_ft.to(device)
    
        return covid_ft, normal_ft, pneumonia_ft, cov_nor_ft, cov_pneu_ft, nor_pneu_ft, criterion
    else: # Previous OVR
        nor_pneu_ft = models.resnet50(pretrained=True)
        nor_pneu_num_ftrs = nor_pneu_ft.fc.in_features
        nor_pneu_ft.fc = nn.Linear(nor_pneu_num_ftrs, 2)
#         nor_pneu_ft.load_state_dict(torch.load(f"{model_dir}/normal_pneumonia_binary.pt"))
        nor_pneu_ft.load_state_dict(torch.load(f"{model_dir}/final_pneumonia_normal_binary.pt"))
        nor_pneu_ft = nor_pneu_ft.to(device)
    
        return covid_ft, nor_pneu_ft, criterion

def get_model_ft(model_dir, device):
    # Create the covid binary model
    covid_ft = models.resnet50(pretrained=True)
    covid_num_ftrs = covid_ft.fc.in_features
    covid_ft.fc = nn.Linear(covid_num_ftrs, 2)
    covid_ft.load_state_dict(torch.load(f"{model_dir}/final_covid_binary.pt"))
    
    # Create the normal binary model
    normal_ft = models.resnet50(pretrained=True)
    normal_num_ftrs = normal_ft.fc.in_features
    normal_ft.fc = nn.Linear(normal_num_ftrs, 2)
    normal_ft.load_state_dict(torch.load(f"{model_dir}/final_normal_binary.pt"))
    
    # Create the pneumonia binary model
    pneumonia_ft = models.resnet50(pretrained=True)
    pneumonia_num_ftrs = pneumonia_ft.fc.in_features
    pneumonia_ft.fc = nn.Linear(pneumonia_num_ftrs, 2)
    pneumonia_ft.load_state_dict(torch.load(f"{model_dir}/final_pneumonia_normal_binary.pt"))
    
    avidnet = AVIDNet(covid_ft, normal_ft, pneumonia_ft)
    avidnet = avidnet.to(device)
    
    # Create the loss function
    criterion = nn.CrossEntropyLoss()
    optimizer = optim.SGD(avidnet.parameters(), lr=0.0001, momentum=0.9)
    lr_scheduler = optim.lr_scheduler.StepLR(optimizer, step_size=3, gamma=0.1)
    
    return avidnet, criterion, optimizer, lr_scheduler

# def train_model(model, dataloaders, criterion, optimizer, lr_scheduler, device, epochs=10):
#     for epoch in range(epochs):
#         model.train()
        
#         epoch_loss = 0.0
#         for inputs, labels in dataloaders:
#             inputs = inputs.to(device)
#             labels = labels.to(device)
            
#             optimizer.zero_grad()
            
#             with torch.set_grad_enabled(True):
#                 outputs = model(inputs)
#                 _, preds = torch.max(outputs, 1)
#                 loss = criterion(outputs, labels)
#                 epoch_loss += loss.item() * inputs.size(0)
                
#                 loss.backward()
#                 optimizer.step()
                
#         lr_scheduler.step()
#         print(f'Epoch: {epoch}/{epochs}, Loss: {epoch_loss}')
        
#     return model

def eval_model(dataloaders=None, covid_model=None, normal_model=None,
               pneumonia_model=None, cov_nor_model=None, cov_pneu_model=None, 
               nor_pneu_model=None, avidnet=None,
               criterion=None, batch_size=1, device=None, out_dir=None,
               mtype=None, confidence_weights=None, covid_threshold=0.95):
    '''Evaluate the model.
    
    Args:
        dataloaders (obj): dataloaders for the evalutation
        covid_model (obj): covid-19 binary model (0: covid-19, 1: non-covid-19)
        normal_model (obj): normal binary model (0: non-normal, 1: normal)
        pneumonia_model (obj): pneumonia binary model (0: non-pneumonia, 1: pneumonia)
        nor_pneu_model (obj): binary model to classify normal and pneumonia
                              (0: normal, 1: pneumonia)
        avidnet (obj): 3-class classification model (0: covid-19, 1: normal, 2: pneumonia)
        criterion (obj): loss function
        batch_size (int): batch size
        device (obj): device to evaluate the model
        out_dir (str): path to save prediction results
        mtype (str): one of the OVR model types
        confidence_weights (list): rate of confidence score for covid-19, normal, pneumonia
    Returns:
        preds_dict (dict): prediction results
    '''
    
    def create_workbook():
        wb = Workbook()      # 워크북을 생성한다.
        ws = wb.active       # 워크 시트를 얻는다.

        ws['A1'] = 'ResNet50'
        ws['B1'] = 'Val ACC'

        ws['D1'] = 'Covid ACC'
        ws['E1'] = 'Covid Recall'
        ws['F1'] = 'Covid Precision'
        ws['G1'] = 'Covid F1'

        ws['I1'] = 'Normal ACC'
        ws['J1'] = 'Normal Recall'
        ws['K1'] = 'Normal Precision'
        ws['L1'] = 'Normal F1'

        ws['N1'] = 'Pneumonia ACC'
        ws['O1'] = 'Pneumonia Recall'
        ws['P1'] = 'Pneumonia Precision'
        ws['Q1'] = 'Pneumonia F1'
        
        return wb, ws
    
    since = time.time()
    
    # Create the workbook and worksheet
    wb, ws = create_workbook()

    # Init parameters
    val_covid_TP, val_normal_TP, val_pneumonia_TP = 0.0, 0.0, 0.0
    val_covid_FN, val_normal_FN, val_pneumonia_FN = 0.0, 0.0, 0.0
    val_covid_TN, val_normal_TN, val_pneumonia_TN = 0.0, 0.0, 0.0
    val_covid_FP, val_normal_FP, val_pneumonia_FP = 0.0, 0.0, 0.0
    
    running_corrects = 0.0
#     covid_threshold = 0.95

    print('Evaluation Result')
    print('-' * 10)

    # Set model to evaluate mode
    if mtype == 'OVR':
        covid_model.eval()
        pneumonia_model.eval()   
        normal_model.eval()   
        cov_nor_model.eval()
        cov_pneu_model.eval()
        nor_pneu_model.eval()
    elif mtype == 'MOVR': # Previous OVR
        covid_model.eval()
        nor_pneu_model.eval()
    else: # AVIDNet-OVR
        avidnet.eval()

    # Iterate over data.
    preds_dict = defaultdict(list)
    for inputs, labels in dataloaders:
        inputs = inputs.to(device)
        labels = labels.to(device)

        correct_label = labels.data[0].item()

        if mtype == 'OVR': # Basic OVR model
            # Predict using covid-19 binary model
            covid_outputs = covid_model(inputs)
            _, covid_preds = torch.max(covid_outputs, 1)
            covid_label = covid_preds.item()
            covid_confidence = torch.nn.functional.softmax(covid_outputs, dim=1)
            covid_confidence_score = covid_confidence[0][0].item()

            # Predict using normal binary model
            normal_outputs = normal_model(inputs)
            _, normal_preds = torch.max(normal_outputs, 1)
            normal_label = normal_preds.item()
            normal_confidence = torch.nn.functional.softmax(normal_outputs, dim=1)
            normal_confidence_score = normal_confidence[0][1].item()

            # Predict using pneumonia binary model
            pneumonia_outputs = pneumonia_model(inputs)
            _, pneumonia_preds = torch.max(pneumonia_outputs, 1)
            pneumonia_label = pneumonia_preds.item()
            pneumonia_confidence = torch.nn.functional.softmax(pneumonia_outputs, dim=1)
            pneumonia_confidence_score = pneumonia_confidence[0][1].item()

            # Calculate total confidence scores
#             covid_confidence_score = normal_confidence[0][0].item()
#             pneumonia_confidence_score = normal_confidence[0][0].item()
            confidence_scores = [covid_confidence_score, normal_confidence_score,
                                 pneumonia_confidence_score]
    
            max_label = np.argmax(confidence_scores)
            th1, th2, th3 = 0.5, 0.9, 0.9
            if True and (max(confidence_scores) < th1 and max_label == 0 \
                    or   max(confidence_scores) < th2 and max_label == 1 \
                    or   max(confidence_scores) < th3 and max_label == 2):
                if True: # OVR
                    # https://kanoki.org/2020/01/14/find-k-smallest-and-largest-values-and-its-indices-in-a-numpy-array/
                    top2_cls = np.argpartition(confidence_scores, -2)[-2:]
                    if 0 in top2_cls and 1 in top2_cls: # covid_normal
                        cov_nor_outputs = cov_nor_model(inputs)
                        _, cov_nor_preds = torch.max(cov_nor_outputs, 1)
                        pre_label = cov_nor_preds.item()
                        # print('Before COVID vs. Normal:', np.argmax(confidence_scores))
                        # print('After COVID vs. Normal:', pre_label)
                    elif 0 in top2_cls and 2 in top2_cls: # covid_pneumonia
                        cov_pneu_outputs = cov_pneu_model(inputs)
                        _, cov_pneu_preds = torch.max(cov_pneu_outputs, 1)
                        pre_label = cov_pneu_preds.item()
                        if pre_label == 1: # pneumonia
                            pre_label += 1
                        # print('Before COVID vs. Pneumonia:', np.argmax(confidence_scores))
                        # print('After COVID vs. Pneumonia:', pre_label)
                    else: # normal_pneumonia
                        nor_pneu_outputs = nor_pneu_model(inputs)
                        _, nor_pneu_preds = torch.max(nor_pneu_outputs, 1)
                        pre_label = nor_pneu_preds.item() + 1
                        # print('Before Normal vs. Pneumonia:', np.argmax(confidence_scores))
                        # print('After Normal vs. Pneumonia:', pre_label)
                else: # OVO
                    count = [0, 0, 0]
                    cov_nor_outputs = cov_nor_model(inputs)
                    _, cov_nor_preds = torch.max(covid_outputs, 1)
                    cov_nor_label = covid_preds.item()
                    count[cov_nor_label] += 1
                        
                    cov_pneu_outputs = cov_pneu_model(inputs)
                    _, cov_pneu_preds = torch.max(cov_pneu_outputs, 1)
                    cov_pneu_label = cov_pneu_preds.item()
                    if cov_pneu_label == 1: # pneumonia
                        cov_pneu_label += 1
                    count[cov_pneu_label] += 1
                        
                    nor_pneu_outputs = cov_pneu_model(inputs)
                    _, nor_pneu_preds = torch.max(nor_pneu_outputs, 1)
                    nor_pneu_label = nor_pneu_preds.item() + 1
                    count[nor_pneu_label] += 1
                    pre_label = np.argmax(count)
#             if confidence_weights:
#                 confidence_scores = np.array(confidence_scores) * np.array(confidence_weights)
            else:
                pre_label = np.argmax(confidence_scores)
        elif mtype == 'MOVR': # Previous OVR model
            # Get the label for covid-19
            covid_outputs = covid_model(inputs)
            _, covid_preds = torch.max(covid_outputs, 1)
            covid_label = covid_preds.item()
            
            # Get the label for normal or pneumonia
            nor_pneu_outputs = nor_pneu_model(inputs)
            _, nor_pneu_preds = torch.max(nor_pneu_outputs, 1)
            binary_label = nor_pneu_preds.item()  #pneumonia_label = 0 = normal/ penumonia_label = 1 = penumonia
            
            # Calculate covid-19 confidence score
            covid_confidence = torch.nn.functional.softmax(covid_outputs, dim=1)
            covid_confidence_score = covid_confidence[0][0].item()
                
            # Calculate normal confidence score
            normal_confidence = torch.nn.functional.softmax(nor_pneu_outputs, dim=1)
            normal_confidence_score = normal_confidence[0][0].item()
            
            # Calculate pneumonia confidence score
            pneumonia_confidence = torch.nn.functional.softmax(nor_pneu_outputs, dim=1)
            pneumonia_confidence_score = pneumonia_confidence[0][1].item()
            
            # Calculate total confidence scores
            confidence_scores = [covid_confidence_score, normal_confidence_score,
                                 pneumonia_confidence_score]
            if confidence_weights:
                confidence_scores = np.array(confidence_scores) * np.array(confidence_weights)

            # Cut off the covid label
            if covid_label == 0 and covid_confidence_score > covid_threshold:
                pre_label = 0
            else:
                confidence_scores[0] = 0.0
                pre_label = np.argmax(confidence_scores)

        else: # Fine-tuning model
            avidnet_outputs = avidnet(inputs)
            _, avidnet_preds = torch.max(avidnet_outputs, 1)
            avidnet_label = avidnet_preds.item()
            
            avidnet_confidence = torch.nn.functional.softmax(avidnet_outputs, dim=1)
            covid_confidence_score = avidnet_confidence[0][0].item()
            normal_confidence_score = avidnet_confidence[0][1].item()
            pneumonia_confidence_score = avidnet_confidence[0][2].item()

            # Calculate total confidence scores
            confidence_scores = [covid_confidence_score, normal_confidence_score,
                                 pneumonia_confidence_score]
            if confidence_weights:
                confidence_scores = np.array(confidence_scores) * np.array(confidence_weights)
            pre_label = np.argmax(confidence_scores)
            
        # Save and print the prediction result
        preds_dict[(pre_label, correct_label)].append(confidence_scores)
#         print(pre_label, correct_label, confidence_scores)
                
        # Calculate metrics
        if pre_label == correct_label: 
            #Covid-index:0, Normal-index:1, Pneumonia-index:2
            running_corrects += 1 
            if pre_label == 0: 
                val_covid_TP += 1 #COVID 관점(Covid->Positive)에서는 covid를 정확히 분류하는 것이 TP
                val_normal_TN += 1 #Normal 관점에서는 covid를 정확히 분류하는 것이 TN
                val_pneumonia_TN += 1 #Pneumonia 관점에서는 covid를 정확히 분류하는 것이 TN

            elif pre_label == 1: 
                val_normal_TP += 1 #Normal 관점에서는 normal를 정확히 분류하는 것이 TP
                val_covid_TN += 1 #COVID 관점에서는 normal를 정확히 분류하는 것이 TN
                val_pneumonia_TN += 1 #Pneumonia 관점에서는 normal를 정확히 분류하는 것이 TN

            elif pre_label == 2: 
                val_pneumonia_TP += 1 #Pneumonia 관점에서는 pneumonia를 정확히 분류하는 것이 TP
                val_covid_TN += 1 #COVID 관점에서는 pneumonia를 정확히 분류하는 것이 TN
                val_normal_TN += 1 #COVID 관점에서는 pneumonia를 정확히 분류하는 것이 TN

        elif pre_label != correct_label:
            if pre_label == 0:
                if correct_label == 1:
                    val_covid_FP += 1 #COVID 관점(Covid->Positive)에서 covid라고 분류했지만 실제로는 normal인 경우는 FP
                    val_normal_FN += 1 #Normal 관점에서 covid라고 분류했지만 실제로는 normal인 경우는 FN
                    val_pneumonia_TN += 1 #<--추가 
                    #TN --> preds:0, label:1   
                    #--> Pneumonia 관점에서는 covid, normal 모두 Negative이다. 
                    #--> 즉, pred:0 --> covid --> Negative/ label:1 --> normal --> Negative
                    #--> 결과적으로 Penumonia관점에서는 실제 Negative(=normal)를 Negative(=covid)라고 분류했기 때문에 True가 된다.
                elif correct_label == 2:
                    val_covid_FP += 1 #COVID 관점에서 covid라고 분류했지만 실제로는 pneumonia인 경우는 FP
                    val_pneumonia_FN += 1 #Pneumonia 관점에서 covid라고 분류했지만 실제로는 pneumonia인 경우는 FN
                    val_normal_TN += 1 #<--추가

            elif pre_label == 1:
                if correct_label == 0:
                    val_normal_FP += 1 #Normal 관점에서 normal이라고 분류했지만 실제로는 covid인 경우는 FP
                    val_covid_FN += 1 #COVID 관점에서 normal이라고 분류했지만 실제로는 covid인 경우는 FN
                    val_pneumonia_TN += 1 #<--추가
                elif correct_label == 2:
                    val_normal_FP += 1 #NORMAL 관점에서 normal이라고 분류했지만 실제로는 pneumonia인 경우는 FP
                    val_pneumonia_FN += 1 #Pneumonia 관점에서 normal라고 분류했지만 실제로는 pneumonia인 경우는 FN
                    val_covid_TN += 1 #<-- 추가

            elif pre_label == 2:
                if correct_label == 0:
                    val_pneumonia_FP += 1 #Pneumonia 관점에서 pneumonia라고 분류했지만 실제로는 covid인 경우는 FP
                    val_covid_FN += 1 #COVID 관점에서 pneumonia라고 분류했지만 실제로는 covid인 경우는 FN
                    val_normal_TN += 1 #<-- 추가
                elif correct_label == 1:
                    val_pneumonia_FP += 1 #Pneumonia 관점에서 pneumonia라고 분류했지만 실제로는 normal인 경우는 FP
                    val_normal_FN += 1 #Normal 관점에서 pneumonia라고 분류했지만 실제로는 normal인 경우는 FN
                    val_covid_TN += 1 #<--추가

    #recall, precision -> https://en.wikipedia.org/wiki/Precision_and_recall
    #recall = TP/(TP+FN)
    #precision = TP/(TP+FP)
    '''
    if val_false_TP + val_false_FP == 0:
        val_false_FP = 0.000001
    '''    
    epsilon = 0.0000001
        
    #Recall
    covid_recall = round(val_covid_TP/(val_covid_TP + val_covid_FN),4)*100
    normal_recall = round(val_normal_TP/(val_normal_TP + val_normal_FN),4)*100
    pneumonia_recall = round(val_pneumonia_TP/(val_pneumonia_TP + val_pneumonia_FN),4)*100

    #Preicision
    covid_precision = round(val_covid_TP/(val_covid_TP + val_covid_FP),4)*100
    normal_precision = round(val_normal_TP/(val_normal_TP + val_normal_FP),4)*100
    pneumonia_precision = round(val_pneumonia_TP/(val_pneumonia_TP + val_pneumonia_FP + epsilon),4)*100
        
    #ACC: Accuracy
    covid_acc = round((val_covid_TP + val_covid_TN)/(val_covid_TP + val_covid_FP +val_covid_TN + val_covid_FN),4)*100
    normal_acc = round((val_normal_TP + val_normal_TN)/(val_normal_TP + val_normal_FP +val_normal_TN + val_normal_FN),4)*100
    pneumonia_acc = round((val_pneumonia_TP + val_pneumonia_TN)/(val_pneumonia_TP + val_pneumonia_FP +val_pneumonia_TN + val_pneumonia_FN),4)*100
        

    '''
    if false_precision+false_recall == 0:
        false_precision = 0.000001
    '''
        
    #F1 score -> https://en.wikipedia.org/wiki/F-score
    #F1 score = 2/((1/recall)+(1/precision)) = 2((precision*recall)/(precision+recall)) = tp/(tp+((1/2)(fp+fn)))
    covid_f1_score = round(2*(covid_precision*covid_recall)/(covid_precision+covid_recall),2)
    normal_f1_score = round(2*(normal_precision*normal_recall)/(normal_precision+normal_recall),2)
    pneumonia_f1_score = round(2*(pneumonia_precision*pneumonia_recall)/(pneumonia_precision+pneumonia_recall + epsilon),2)

    #total_data = batch size --> covid_total_dataset = false_total_dataset
    covid_total_dataset = int(val_covid_TP + val_covid_FP +val_covid_TN + val_covid_FN)
    normal_total_dataset = int(val_normal_TP + val_normal_FP +val_normal_TN + val_normal_FN)
    pneumonia_total_dataset = int(val_pneumonia_TP + val_pneumonia_FP +val_pneumonia_TN + val_pneumonia_FN)
        
    total_acc = running_corrects / covid_total_dataset
    
    # Write results to the work sheet
    ws['B2'] = total_acc
    
    ws['D2'] = covid_acc
    ws['E2'] = covid_recall
    ws['F2'] = covid_precision
    ws['G2'] = covid_f1_score
        
    ws['I2'] = normal_acc
    ws['J2'] = normal_recall
    ws['K2'] = normal_precision
    ws['L2'] = normal_f1_score

    ws['N2'] = pneumonia_acc
    ws['O2'] = pneumonia_recall
    ws['P2'] = pneumonia_precision
    ws['Q2'] = pneumonia_f1_score

      
    print('Val Acc: {:.4f}'.format(total_acc))    

    print()
    print('데이터셋A: ' + str(covid_total_dataset))
    print('Covid Recall: ' + str(covid_recall))        
    print('Covid Precision: ' + str(covid_precision))
    print('Covid ACC: ' + str(covid_acc))
    print('Covid F1 score: ' + str(covid_f1_score))
    print()
        
    print('데이터셋B: ' + str(normal_total_dataset))
    print('Normal Recall: ' + str(normal_recall))
    print('Normal Precision: ' + str(normal_precision))
    print('Normal ACC: ' + str(normal_acc))
    print('Normal F1 score: ' + str(normal_f1_score))
    print()
        
    print('데이터셋C: ' + str(pneumonia_total_dataset))
    print('Pneumonia Recall: ' + str(pneumonia_recall))
    print('Pneumonia Precision: ' + str(pneumonia_precision))
    print('Pneumonia ACC: ' + str(pneumonia_acc))
    print('Pneumonia F1 score: ' + str(pneumonia_f1_score))
        
    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    #print('Best val Acc: {:4f}'.format(best_acc))

    # load best model weights
    #model.load_state_dict(best_covid_wts)
    if not os.path.isdir(out_dir):
        os.mkdir(out_dir)
        
    wb.save(f'{out_dir}/integrate_test.xlsx') # 엑셀로 저장한다. 
    #torch.save(model.state_dict(), 'covid_binary.pt')
    
    return preds_dict, total_acc
    
    
# class AVIDNet(nn.Module):
#     '''https://discuss.pytorch.org/t/custom-ensemble-approach/52024/4'''
    
#     def __init__(self, modelA, modelB, modelC, nb_classes=3):
#         super(AVIDNet, self).__init__()
#         self.modelA = modelA
#         self.modelB = modelB
#         self.modelC = modelC
        
#         # Remove last linear layer
#         num_ftrs = self.modelA.fc.in_features
#         self.modelA.fc = nn.Identity()
#         self.modelB.fc = nn.Identity()
#         self.modelC.fc = nn.Identity()
        
#         # Create new classifier
#         self.classifier = nn.Linear(num_ftrs*3, nb_classes)
        
#     def forward(self, x):
#         x1 = self.modelA(x.clone())  # clone to make sure x is not changed by inplace methods
#         x1 = x1.view(x1.size(0), -1)
#         x2 = self.modelB(x.clone())
#         x2 = x2.view(x2.size(0), -1)
#         x3 = self.modelC(x)
#         x3 = x3.view(x3.size(0), -1)
#         x = torch.cat((x1, x2, x3), dim=1)
        
#         x = self.classifier(F.relu(x))
#         return x
